import csv
import io
import logging
import zipfile
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_supabase_client
from app.services.grading_service import get_grading_service

logger = logging.getLogger(__name__)


# ── Styling constants ────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_GREEN_FONT = Font(name="Calibri", color="006100")
_YELLOW_FONT = Font(name="Calibri", color="9C6500")
_RED_FONT = Font(name="Calibri", color="9C0006")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


# ── Helpers ──────────────────────────────────────────────────

def _perf_level(pct: float) -> str:
    if pct >= 80:
        return "Good"
    elif pct >= 60:
        return "Average"
    return "Needs Improvement"


def _perf_style(level: str) -> tuple[PatternFill, Font]:
    if level == "Good":
        return _GREEN_FILL, _GREEN_FONT
    elif level == "Average":
        return _YELLOW_FILL, _YELLOW_FONT
    return _RED_FILL, _RED_FONT


def _style_header_row(ws: Any, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws: Any, min_width: int = 10, max_width: int = 45) -> None:
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = length


# ── ExportService ────────────────────────────────────────────

class ExportService:
    """Service for exporting session data to Excel, CSV, and PDF."""

    # ── Data fetching ────────────────────────────────────────

    async def _fetch_session_data(
        self, session_id: str, teacher_id: str
    ) -> dict[str, Any]:
        """Fetch all data needed for export: session, template, exams, results."""
        sb = await get_supabase_client()

        # Session
        session_res = (
            await sb.table("grading_sessions")
            .select("*")
            .eq("id", session_id)
            .eq("teacher_id", teacher_id)
            .limit(1)
            .execute()
        )
        if not session_res.data:
            raise ValueError(f"Session '{session_id}' not found")
        session = session_res.data[0]

        # Template
        tmpl_res = (
            await sb.table("exam_templates")
            .select("*")
            .eq("id", session["template_id"])
            .limit(1)
            .execute()
        )
        template = tmpl_res.data[0] if tmpl_res.data else {}

        # Student exams
        exams_res = (
            await sb.table("student_exams")
            .select("*")
            .eq("session_id", session_id)
            .order("created_at", desc=False)
            .execute()
        )
        exams = exams_res.data or []

        # Grading results
        exam_ids = [e["id"] for e in exams]
        results: list[dict[str, Any]] = []
        if exam_ids:
            results_res = (
                await sb.table("grading_results")
                .select("*")
                .in_("exam_id", exam_ids)
                .execute()
            )
            results = results_res.data or []

        results_map = {r["exam_id"]: r for r in results}

        return {
            "session": session,
            "template": template,
            "exams": exams,
            "results": results,
            "results_map": results_map,
        }

    # ── 1. Excel export ──────────────────────────────────────

    async def export_session_to_excel(
        self, session_id: str, teacher_id: str
    ) -> bytes:
        """Export session data to a multi-sheet Excel workbook."""
        logger.info("Exporting session %s to Excel", session_id)

        data = await self._fetch_session_data(session_id, teacher_id)
        session = data["session"]
        template = data["template"]
        exams = data["exams"]
        results_map = data["results_map"]

        max_score = float(template.get("max_score", 0))
        structure = template.get("structure_json", {})
        grading = get_grading_service()

        # Precompute per-student data
        student_data: list[dict[str, Any]] = []
        for exam in exams:
            result = results_map.get(exam["id"])
            total = float(result["total_score"]) if result else 0
            final = float(result["final_score"]) if result and result.get("final_score") is not None else total
            pct = (final / max_score * 100) if max_score > 0 else 0
            level = _perf_level(pct)
            section_scores = result.get("section_scores_json", {}) if result else {}

            student_data.append({
                "exam": exam,
                "result": result,
                "name": exam.get("student_name") or "Unknown",
                "total": total,
                "final": final,
                "max": max_score,
                "pct": pct,
                "level": level,
                "status": exam.get("status", "pending"),
                "section_scores": section_scores,
            })

        # Generate improvement plans for graded students
        plans: dict[str, dict[str, Any]] = {}
        for sd in student_data:
            if sd["result"] and sd["status"] == "graded":
                try:
                    plan = await grading.generate_improvement_plan(
                        student_results=sd["result"],
                        template=structure,
                        student_name=sd["name"],
                    )
                    plans[sd["name"]] = plan
                except Exception as exc:
                    logger.warning("Failed to generate plan for '%s': %s", sd["name"], exc)

        wb = Workbook()

        # ── Sheet 1: Summary ─────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"

        scores = [sd["final"] for sd in student_data if sd["status"] == "graded"]
        pass_count = sum(1 for sd in student_data if sd["pct"] >= 60)
        graded_count = len(scores)

        summary_rows = [
            ("Session Name", session.get("name", "")),
            ("Date", session.get("created_at", "")[:10] if session.get("created_at") else ""),
            ("Template", template.get("name", "")),
            ("Subject", template.get("subject", "")),
            ("Max Score", max_score),
            ("Total Students", len(exams)),
            ("Graded", graded_count),
            ("", ""),
            ("CLASS STATISTICS", ""),
            ("Average Score", f"{sum(scores) / len(scores):.1f}" if scores else "N/A"),
            ("Highest Score", f"{max(scores):.1f}" if scores else "N/A"),
            ("Lowest Score", f"{min(scores):.1f}" if scores else "N/A"),
            ("Pass Rate (≥60%)", f"{pass_count}/{graded_count} ({pass_count / graded_count * 100:.0f}%)" if graded_count else "N/A"),
        ]

        # Common weaknesses from plans
        all_weaknesses: list[str] = []
        for plan in plans.values():
            all_weaknesses.extend(plan.get("score_analysis", {}).get("weaknesses", []))
        if all_weaknesses:
            from collections import Counter
            common = Counter(all_weaknesses).most_common(5)
            summary_rows.append(("", ""))
            summary_rows.append(("COMMON WEAKNESSES", ""))
            for weakness, count in common:
                summary_rows.append(("", f"{weakness} ({count} students)"))

        for row_data in summary_rows:
            ws_summary.append(row_data)

        # Style summary
        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, max_col=2):
            row[0].font = Font(name="Calibri", bold=True, size=11)
            row[1].font = Font(name="Calibri", size=11)
        ws_summary.column_dimensions["A"].width = 22
        ws_summary.column_dimensions["B"].width = 50

        # ── Sheet 2: Grades ──────────────────────────────────
        ws_grades = wb.create_sheet("Grades")
        headers = ["Student Name", "Total Score", "Max Score", "%", "Status", "Performance"]
        ws_grades.append(headers)
        _style_header_row(ws_grades, len(headers))

        for i, sd in enumerate(student_data, start=2):
            ws_grades.append([
                sd["name"],
                round(sd["final"], 2),
                sd["max"],
                round(sd["pct"], 1),
                sd["status"],
                sd["level"],
            ])
            # Color-code performance
            fill, font = _perf_style(sd["level"])
            perf_cell = ws_grades.cell(row=i, column=6)
            perf_cell.fill = fill
            perf_cell.font = font
            # Color-code percentage
            pct_cell = ws_grades.cell(row=i, column=4)
            pct_cell.fill = fill
            pct_cell.font = font

            for col in range(1, len(headers) + 1):
                ws_grades.cell(row=i, column=col).border = _THIN_BORDER

        _auto_width(ws_grades)

        # ── Sheet 3: Section Breakdown ───────────────────────
        ws_sections = wb.create_sheet("Section Breakdown")
        sec_headers = ["Student", "Section", "Points", "Max", "%", "Status"]
        ws_sections.append(sec_headers)
        _style_header_row(ws_sections, len(sec_headers))

        row_num = 2
        for sd in student_data:
            for section_name, scores_data in sd["section_scores"].items():
                earned = float(scores_data.get("earned", 0))
                sec_max = float(scores_data.get("max", 0))
                sec_pct = (earned / sec_max * 100) if sec_max > 0 else 0
                sec_status = "Strong" if sec_pct >= 70 else ("Average" if sec_pct >= 50 else "Weak")

                ws_sections.append([
                    sd["name"], section_name, round(earned, 2),
                    round(sec_max, 2), round(sec_pct, 1), sec_status,
                ])

                fill, font = _perf_style(
                    "Good" if sec_pct >= 70 else ("Average" if sec_pct >= 50 else "Needs Improvement")
                )
                ws_sections.cell(row=row_num, column=6).fill = fill
                ws_sections.cell(row=row_num, column=6).font = font
                ws_sections.cell(row=row_num, column=5).fill = fill
                ws_sections.cell(row=row_num, column=5).font = font

                for col in range(1, len(sec_headers) + 1):
                    ws_sections.cell(row=row_num, column=col).border = _THIN_BORDER
                row_num += 1

        _auto_width(ws_sections)

        # ── Sheet 4: Individual Plans ────────────────────────
        ws_plans = wb.create_sheet("Individual Plans")
        plan_headers = [
            "Student", "Overall Performance", "Top Strengths",
            "Main Weaknesses", "Priority Actions",
        ]
        ws_plans.append(plan_headers)
        _style_header_row(ws_plans, len(plan_headers))

        row_num = 2
        for sd in student_data:
            plan = plans.get(sd["name"], {})
            analysis = plan.get("score_analysis", {})
            impr = plan.get("improvement_plan", {})
            actions = impr.get("immediate_actions", [])

            strengths_str = "\n".join(analysis.get("strengths", ["-"]))
            weaknesses_str = "\n".join(analysis.get("weaknesses", ["-"]))
            actions_str = "\n".join(
                f"• {a.get('topic', '')}: {a.get('description', '')}"
                for a in actions
            ) or "-"

            ws_plans.append([
                sd["name"],
                plan.get("overall_performance", "-"),
                strengths_str,
                weaknesses_str,
                actions_str,
            ])

            fill, font = _perf_style(plan.get("overall_performance", "Average"))
            ws_plans.cell(row=row_num, column=2).fill = fill
            ws_plans.cell(row=row_num, column=2).font = font

            for col in range(1, len(plan_headers) + 1):
                ws_plans.cell(row=row_num, column=col).border = _THIN_BORDER
                ws_plans.cell(row=row_num, column=col).alignment = _WRAP_ALIGN
            row_num += 1

        _auto_width(ws_plans, max_width=60)

        # ── Sheet 5: Detailed Plans ──────────────────────────
        ws_detailed = wb.create_sheet("Detailed Plans")
        ws_detailed.column_dimensions["A"].width = 18
        ws_detailed.column_dimensions["B"].width = 30
        ws_detailed.column_dimensions["C"].width = 45
        ws_detailed.column_dimensions["D"].width = 35
        ws_detailed.column_dimensions["E"].width = 18

        row_num = 1
        for sd in student_data:
            plan = plans.get(sd["name"])
            if not plan:
                continue

            # Student header
            ws_detailed.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num, end_column=5,
            )
            header_cell = ws_detailed.cell(row=row_num, column=1)
            header_cell.value = f"{sd['name']} — {sd['final']}/{sd['max']} ({sd['pct']:.0f}%)"
            header_cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center")
            row_num += 1

            # Performance analysis
            ws_detailed.cell(row=row_num, column=1, value="Overall Performance")
            ws_detailed.cell(row=row_num, column=1).font = _SUBHEADER_FONT
            perf = plan.get("overall_performance", "-")
            perf_cell = ws_detailed.cell(row=row_num, column=2, value=perf)
            fill, font = _perf_style(perf)
            perf_cell.fill = fill
            perf_cell.font = font
            row_num += 1

            # Strengths
            ws_detailed.cell(row=row_num, column=1, value="Strengths")
            ws_detailed.cell(row=row_num, column=1).font = _SUBHEADER_FONT
            for s in plan.get("score_analysis", {}).get("strengths", []):
                ws_detailed.cell(row=row_num, column=2, value=s)
                row_num += 1
            if not plan.get("score_analysis", {}).get("strengths"):
                row_num += 1

            # Weaknesses
            ws_detailed.cell(row=row_num, column=1, value="Weaknesses")
            ws_detailed.cell(row=row_num, column=1).font = _SUBHEADER_FONT
            for w in plan.get("score_analysis", {}).get("weaknesses", []):
                ws_detailed.cell(row=row_num, column=2, value=w)
                row_num += 1
            if not plan.get("score_analysis", {}).get("weaknesses"):
                row_num += 1

            # Action tables for each plan tier
            for tier_key, tier_label in [
                ("immediate_actions", "Immediate Actions"),
                ("medium_term_goals", "Medium-Term Goals"),
                ("long_term_goals", "Long-Term Goals"),
            ]:
                items = plan.get("improvement_plan", {}).get(tier_key, [])
                if not items:
                    continue

                # Tier header
                ws_detailed.merge_cells(
                    start_row=row_num, start_column=1,
                    end_row=row_num, end_column=5,
                )
                tier_cell = ws_detailed.cell(row=row_num, column=1, value=tier_label)
                tier_cell.font = _SUBHEADER_FONT
                tier_cell.fill = _SUBHEADER_FILL
                row_num += 1

                # Sub-headers
                for ci, h in enumerate(["Topic", "Description", "Resources", "Est. Time"], start=1):
                    c = ws_detailed.cell(row=row_num, column=ci, value=h)
                    c.font = Font(name="Calibri", bold=True, size=10)
                    c.border = _THIN_BORDER
                row_num += 1

                for item in items:
                    resources = ", ".join(item.get("resources", []))
                    ws_detailed.cell(row=row_num, column=1, value=item.get("topic", ""))
                    ws_detailed.cell(row=row_num, column=2, value=item.get("description", ""))
                    ws_detailed.cell(row=row_num, column=3, value=resources)
                    ws_detailed.cell(row=row_num, column=4, value=item.get("estimated_time", ""))
                    for ci in range(1, 5):
                        ws_detailed.cell(row=row_num, column=ci).border = _THIN_BORDER
                        ws_detailed.cell(row=row_num, column=ci).alignment = _WRAP_ALIGN
                    row_num += 1

            # Specific recommendations
            recs = plan.get("specific_recommendations", [])
            if recs:
                ws_detailed.cell(row=row_num, column=1, value="Recommendations")
                ws_detailed.cell(row=row_num, column=1).font = _SUBHEADER_FONT
                for rec in recs:
                    ws_detailed.cell(row=row_num, column=2, value=f"• {rec}")
                    ws_detailed.cell(row=row_num, column=2).alignment = _WRAP_ALIGN
                    row_num += 1

            # Next steps
            next_steps = plan.get("next_steps", "")
            if next_steps:
                ws_detailed.cell(row=row_num, column=1, value="Next Steps")
                ws_detailed.cell(row=row_num, column=1).font = _SUBHEADER_FONT
                ws_detailed.cell(row=row_num, column=2, value=next_steps)
                ws_detailed.cell(row=row_num, column=2).alignment = _WRAP_ALIGN
                row_num += 1

            # Spacer between students
            row_num += 2

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info("Excel export complete for session %s (%d bytes)", session_id, buf.getbuffer().nbytes)
        return buf.getvalue()

    # ── 2. CSV export ────────────────────────────────────────

    async def export_session_to_csv(
        self, session_id: str, teacher_id: str
    ) -> str:
        """Export session grades as a simple CSV string."""
        logger.info("Exporting session %s to CSV", session_id)

        data = await self._fetch_session_data(session_id, teacher_id)
        template = data["template"]
        exams = data["exams"]
        results_map = data["results_map"]
        max_score = float(template.get("max_score", 0))

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Student Name", "Total Score", "Max Score", "Percentage", "Status", "Performance Level"])

        for exam in exams:
            result = results_map.get(exam["id"])
            total = float(result["total_score"]) if result else 0
            final = float(result["final_score"]) if result and result.get("final_score") is not None else total
            pct = (final / max_score * 100) if max_score > 0 else 0
            level = _perf_level(pct)

            writer.writerow([
                exam.get("student_name") or "Unknown",
                round(final, 2),
                max_score,
                round(pct, 1),
                exam.get("status", "pending"),
                level,
            ])

        csv_str = output.getvalue()
        logger.info("CSV export complete for session %s (%d chars)", session_id, len(csv_str))
        return csv_str

    # ── 3. Individual PDF export ─────────────────────────────

    async def export_individual_plan_pdf(
        self, exam_id: str, teacher_id: str
    ) -> bytes:
        """Generate a PDF improvement plan for a single student exam."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        logger.info("Generating PDF plan for exam %s", exam_id)

        sb = await get_supabase_client()

        # Fetch exam
        exam_res = (
            await sb.table("student_exams")
            .select("*")
            .eq("id", exam_id)
            .limit(1)
            .execute()
        )
        if not exam_res.data:
            raise ValueError(f"Exam '{exam_id}' not found")
        exam = exam_res.data[0]

        # Verify ownership via session
        session_res = (
            await sb.table("grading_sessions")
            .select("*, exam_templates:template_id(*)")
            .eq("id", exam["session_id"])
            .eq("teacher_id", teacher_id)
            .limit(1)
            .execute()
        )
        if not session_res.data:
            raise ValueError(f"Exam '{exam_id}' not found or access denied")
        session = session_res.data[0]

        # Fetch template separately (in case join doesn't work)
        tmpl_res = (
            await sb.table("exam_templates")
            .select("*")
            .eq("id", session["template_id"])
            .limit(1)
            .execute()
        )
        template = tmpl_res.data[0] if tmpl_res.data else {}
        structure = template.get("structure_json", {})
        max_score = float(template.get("max_score", 0))

        # Fetch result
        result_res = (
            await sb.table("grading_results")
            .select("*")
            .eq("exam_id", exam_id)
            .limit(1)
            .execute()
        )
        result = result_res.data[0] if result_res.data else {}

        total = float(result.get("total_score", 0))
        final = float(result["final_score"]) if result.get("final_score") is not None else total
        pct = (final / max_score * 100) if max_score > 0 else 0
        student_name = exam.get("student_name") or "Unknown Student"

        # Generate improvement plan
        grading = get_grading_service()
        plan = await grading.generate_improvement_plan(
            student_results=result,
            template=structure,
            student_name=student_name,
        )

        # Build PDF
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=letter,
            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
            topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"],
            fontSize=18, spaceAfter=6, textColor=colors.HexColor("#1F3864"),
        )
        heading_style = ParagraphStyle(
            "CustomHeading", parent=styles["Heading2"],
            fontSize=13, spaceAfter=6, spaceBefore=14,
            textColor=colors.HexColor("#2F5496"),
        )
        body_style = ParagraphStyle(
            "CustomBody", parent=styles["Normal"],
            fontSize=10, spaceAfter=4, leading=14,
        )
        bullet_style = ParagraphStyle(
            "CustomBullet", parent=body_style,
            leftIndent=20, bulletIndent=10,
        )

        elements: list[Any] = []

        # Header
        elements.append(Paragraph("Improvement Plan", title_style))
        elements.append(Paragraph(
            f"<b>{student_name}</b> — {template.get('name', '')} ({template.get('subject', '')})",
            body_style,
        ))
        elements.append(Spacer(1, 6))

        # Score summary table
        perf = plan.get("overall_performance", "-")
        score_data = [
            ["Score", "Max", "Percentage", "Performance"],
            [f"{final:.1f}", f"{max_score:.1f}", f"{pct:.1f}%", perf],
        ]
        score_table = Table(score_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 2 * inch])

        perf_color = (
            colors.HexColor("#C6EFCE") if perf == "Good"
            else colors.HexColor("#FFEB9C") if perf == "Average"
            else colors.HexColor("#FFC7CE")
        )
        score_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (3, 1), (3, 1), perf_color),
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 12))

        # Strengths & Weaknesses
        elements.append(Paragraph("Score Analysis", heading_style))
        analysis = plan.get("score_analysis", {})
        for s in analysis.get("strengths", []):
            elements.append(Paragraph(f"<b>✓</b> {s}", bullet_style))
        for w in analysis.get("weaknesses", []):
            elements.append(Paragraph(f"<b>✗</b> {w}", bullet_style))

        # Action plan tiers
        for tier_key, tier_label in [
            ("immediate_actions", "Immediate Actions"),
            ("medium_term_goals", "Medium-Term Goals"),
            ("long_term_goals", "Long-Term Goals"),
        ]:
            items = plan.get("improvement_plan", {}).get(tier_key, [])
            if not items:
                continue

            elements.append(Paragraph(tier_label, heading_style))

            table_data = [["Topic", "Description", "Resources", "Time"]]
            for item in items:
                table_data.append([
                    item.get("topic", ""),
                    item.get("description", ""),
                    ", ".join(item.get("resources", [])),
                    item.get("estimated_time", ""),
                ])

            t = Table(table_data, colWidths=[1.3 * inch, 2.5 * inch, 2 * inch, 1 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D6E4F0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)

        # Specific recommendations
        recs = plan.get("specific_recommendations", [])
        if recs:
            elements.append(Paragraph("Specific Recommendations", heading_style))
            for rec in recs:
                elements.append(Paragraph(f"• {rec}", bullet_style))

        # Next steps
        next_steps = plan.get("next_steps", "")
        if next_steps:
            elements.append(Paragraph("Next Steps", heading_style))
            elements.append(Paragraph(next_steps, body_style))

        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            "Footer", parent=body_style,
            fontSize=8, textColor=colors.grey,
        )
        elements.append(Paragraph(
            f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} — Exam Grader",
            footer_style,
        ))

        doc.build(elements)
        buf.seek(0)
        pdf_bytes = buf.getvalue()
        logger.info("PDF export complete for exam %s (%d bytes)", exam_id, len(pdf_bytes))
        return pdf_bytes

    # ── 4. Bulk PDF export (ZIP) ─────────────────────────────

    async def export_session_pdfs_zip(
        self, session_id: str, teacher_id: str
    ) -> bytes:
        """Generate a ZIP with one PDF per graded student in the session."""
        logger.info("Generating PDF ZIP for session %s", session_id)

        data = await self._fetch_session_data(session_id, teacher_id)
        exams = data["exams"]

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for exam in exams:
                if exam.get("status") != "graded":
                    continue
                try:
                    pdf_bytes = await self.export_individual_plan_pdf(exam["id"], teacher_id)
                    name = (exam.get("student_name") or "unknown").replace(" ", "_")
                    zf.writestr(f"{name}_improvement_plan.pdf", pdf_bytes)
                except Exception as exc:
                    logger.error("Failed to generate PDF for exam %s: %s", exam["id"], exc)

        buf.seek(0)
        zip_bytes = buf.getvalue()
        logger.info("ZIP export complete for session %s (%d bytes)", session_id, len(zip_bytes))
        return zip_bytes


# ── Module-level singleton ───────────────────────────────────

_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
